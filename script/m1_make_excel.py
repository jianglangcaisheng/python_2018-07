
import openpyxl as xls
import os, sys

wb = xls.Workbook()
ws = wb.active

if 0:
    ws['A1'] = 3
    d = ws.cell(row=2, column=3)
    d.value = 5

path_files = input("粘贴路径 & 回车\n")
if 1:
    file_dir = path_files
elif 0:
    file_dir = r"U:\0 image_1809_1_cali_only\0"
elif 0:
    file_dir = os.getcwd()

if 1:
    for root, dirs, files in os.walk(file_dir):

        print("共计%d个文件。" % files.__len__())

        for i_file in range(files.__len__()):
            d = ws.cell(row=i_file + 1, column=1)
            d.value = files[i_file]

wb.save("filename.xlsx")




